import os
import re
from openpyxl import Workbook, load_workbook
import cmath


def read_every_txt(fn, s, datas):
    # 验证数据是否重复的数组
    diff = []
    # 正常数据数
    num_normal = 0
    # 重复数据
    num_repeat = 0
    # 异常数据
    num_error = 0
    num_defect = 0
    i = 0
    while i < len(datas):
        arr = re.findall(r"\d+", datas[i])
        t = arr[0]
        a_list = [arr[3]]
        # a_list.append()
        check_list = [arr[4]]
        check_no = arr[5]
        serial = arr[6]
        # 4个一组
        for j in range(1, 4):
            if i+j > len(datas)-1:
                break
            arr_temp = re.findall(r"\d+", datas[i+j])
            if t != arr_temp[0]:
                # i += j
                # num_error += j
                break
            a_list.append(arr_temp[3])
            check_list.append(arr_temp[4])
        if len(a_list) == 4:
            # for j in range(4):
            # if a_list[0] == check_list[0] and a_list[1] == check_list[1] and \
            #         a_list[2] == check_list[2] and a_list[3] == check_list[3]:
            # if a_list == check_list and a_list not in diff:
            #     li = [t, a_list[0], a_list[1], a_list[2], a_list[3], check_no, serial]
            #     num_normal += 1
            #     diff.append(a_list)
            #     s.append(li)
            # else:

            if a_list != check_list:
                i += 4
                print(filename + "中，Tag标识: " + t + "的数据与校验值不同！")
                break
            # 异常过滤
            # print(str(float(max(a_list)))+"      "+
            #       str(float(min(a_list)))+"        "+str(float(max(a_list))/float(min(a_list))))
            # if float(min([a_list[0], a_list[1], a_list[2], a_list[3]]))/float(max([a_list[0], a_list[1], a_list[2], a_list[3]])) > 10:
            #     num_error += 1
            #     break
            if a_list not in diff:
                li = [t, a_list[0], a_list[1], a_list[2], a_list[3], check_no, serial]
                if float(min(li[1:5]))/float(max(li[1:5])) > 10:
                    num_error += 1
                    i += 4
                    continue
                num_normal += 1
                diff.append([a_list[0], a_list[1], a_list[2], a_list[3]])
                s.append(li)
            else:
                num_repeat += 1
            i += 4
        else:
            num_defect += len(a_list)
            # print(str(i)+ "         ", end='')
            i += len(a_list)
    print(fn+"中共有"+str(len(datas))+"条数据，去除"+str(num_defect)+"缺失数据，"
          +str(num_repeat)+"条重复数据，"+str(num_error)+"条异常数据，经处理后，保留了"+str(num_normal)+"个样本")
    return [num_repeat, num_error, num_normal, num_defect]
    # for data in datas:
    #     # print(data)
    #     arr = re.findall(r"\d+", data)
    #     t = arr[0]
    #     rr = arr[2]
    #     a = arr[3]
    #     check = arr[4]
    #     serial = arr[5]
    #     no = arr[6]
    #     li = [t, rr, a, check, serial]
    #     s.append(li)
        # print(t+"  rr: "+rr+"   a   "+a)


if __name__ == '__main__':
    # (cmath.sqrt(2)*5000)**2+
    root_dir = "error\\"
    all = 0
    all_normal = 0
    all_error = 0
    # 缺失数据
    all_defect = 0
    all_repeat = 0

    # write_path = open('要写入的txt文件的绝对路径/write.txt', 'w')
    for (dirpath, dirnames, filenames) in os.walk(root_dir):
        for filename in filenames:
            # print(filename)
            with open(root_dir+filename, 'r') as txtData:  # 打开文件夹中的文件内容
                # print(fp)
                lines = txtData.readlines()
                # print(lines[1:])
                # 每个txt文件
                wb = Workbook()
                ws = wb.create_sheet()
                sheet = wb.active
                title = ["Tag标识", "A0", "A1", "A2", "A3", "数据序列号", "数据编号"]
                sheet.append(title)
                all += len(lines[1:])
                temp = read_every_txt(filename, sheet, lines[1:])
                all_repeat += temp[0]
                all_error += temp[1]
                all_normal += temp[2]
                all_defect += temp[3]
                # 处理文件名
                name = re.search(r"\d+", filename).group()
                # wb.save("异常数据\\"+name+".异常.xlsx")
            # break
            # print(filename+".正常.xlsx已完成！")
    print("异常数据原有一共" + str(all) + "个样本，一共去除" + str(all_error) + "条异常数据，" + str(all_defect) + "条缺失数据，"
          + str(all_repeat) + "条相同数据，共剩余" + str(all_normal) + "条有效数据。")
    # write_path.close()

