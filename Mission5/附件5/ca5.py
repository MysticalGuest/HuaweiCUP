import re
import openpyxl
# from data_process import extract_num, handle_tag
from openpyxl.styles import colors, PatternFill
import numpy as np
from openpyxl import Workbook, load_workbook
from math import sqrt

from scipy.optimize import fsolve


if __name__ == '__main__':
    # root_dir = "..\\"
    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    title = ["Tag编号", "A0训练值", "A1训练值", "A2训练值", "A3训练值", "A0A1A2x", "A0A1A2y", "A0A1A2z",]
    sheet.append(title)

    wb = openpyxl.load_workbook("附件5有干扰数据预测距离.xlsx", data_only=True)
    # wb = openpyxl.load_workbook("附件5无干扰数据预测距离.xlsx", data_only=True)
    # wb = openpyxl.load_workbook(root_dir + "", data_only=True)
    ws = wb.active
    rows = ws.rows
    # 全部数据
    li = list(rows)
    # i=1
    for ld in li[1:]:
        s0=ld[1].value
        s1 = ld[2].value
        s2 = ld[3].value
        s3 = ld[4].value
        print(s0, s1, s2)
        if s0 is None:
            break
        # print(s0, s1, s2, s3)
        def solve_function123(unsolved_value):
            x, y, z = unsolved_value[0], unsolved_value[1], unsolved_value[2]
            return [
                x ** 2 + y ** 2 + (z - 1300) ** 2 - s0 ** 2,
                (x - 5000) ** 2 + y ** 2 + (z - 1700) ** 2 - s1 ** 2,
                x ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - s2 ** 2,
                # (x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - s3 ** 2,
            ]

        solved123 = fsolve(solve_function123, [0, 0, 0])

        # def solve_function124(unsolved_value):
        #     x, y, z = unsolved_value[0], unsolved_value[1], unsolved_value[2]
        #     return [
        #         x ** 2 + y ** 2 + (z - 1300) ** 2 - s0 ** 2,
        #         (x - 5000) ** 2 + y ** 2 + (z - 1700) ** 2 - s1 ** 2,
        #         (x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - s3 ** 2,
        #     ]
        #
        # solved124 = fsolve(solve_function124, [0, 0, 0])
        #
        # def solve_function134(unsolved_value):
        #     x, y, z = unsolved_value[0], unsolved_value[1], unsolved_value[2]
        #     return [
        #         x ** 2 + y ** 2 + (z - 1300) ** 2 - s0 ** 2,
        #         x ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - s2 ** 2,
        #         (x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - s3 ** 2,
        #     ]
        #
        # solved134 = fsolve(solve_function134, [0, 0, 0])
        #
        # def solve_function234(unsolved_value):
        #     x, y, z = unsolved_value[0], unsolved_value[1], unsolved_value[2]
        #     return [
        #         (x - 5000) ** 2 + y ** 2 + (z - 1700) ** 2 - s1 ** 2,
        #         x ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - s2 ** 2,
        #         (x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - s3 ** 2,
        #     ]

        # solved234 = fsolve(solve_function234, [0, 0, 0])
        sheet.append([int(ld[0].value), s0, s1, s2, s3, solved123[0], solved123[1], solved123[2]])
        # i += 1
    # res_wb.save("根据附件5无干扰训练数据得到坐标.xlsx")
    res_wb.save("根据附件5有干扰训练数据得到坐标.xlsx")