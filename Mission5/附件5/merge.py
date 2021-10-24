import re
import openpyxl
# from data_process import extract_num, handle_tag
from openpyxl.styles import colors, PatternFill
import numpy as np
from openpyxl import Workbook, load_workbook
from math import sqrt

res_wb = Workbook()
res_ws = res_wb.create_sheet()
sheet = res_wb.active
title = ["Tag编号", "A0训练值", "A1训练值", "A2训练值", "A3训练值", "A0A1A2x", "A0A1A2y", "A0A1A2z", "预测精准度",
         "最优解x", "最优解y", "最优解z"]
sheet.append(title)

wb = openpyxl.load_workbook("利用SLSQP算法求得附件5无干扰数据最优解.xlsx", data_only=True)
# wb = openpyxl.load_workbook("根据附件3有干扰训练数据得到坐标.xlsx", data_only=True)
ws = wb.active
rows = ws.rows
# 全部数据
li = list(rows)

for ld in li[1:]:
    print([ld[0].value, ld[1].value, ld[2].value, ld[3].value, ld[4].value,
                  ld[5].value, ld[6].value, ld[7].value, ld[8].value, ld[9].value,
                  ld[10].value, ld[11].value])
    sheet.append([ld[0].value, ld[1].value, ld[2].value, ld[3].value, ld[4].value,
                  ld[5].value, ld[6].value, ld[7].value, ld[8].value, ld[9].value,
                  ld[10].value, ld[11].value])

wb1 = openpyxl.load_workbook("利用SLSQP算法求得附件5有干扰数据最优解.xlsx", data_only=True)
# wb = openpyxl.load_workbook("根据附件3有干扰训练数据得到坐标.xlsx", data_only=True)
ws1 = wb1.active
rows1 = ws1.rows
# 全部数据
li1 = list(rows1)

for ld1 in li1[1:]:
    print([ld1[0].value, ld1[1].value, ld1[2].value, ld1[3].value, ld1[4].value,
           ld1[5].value, ld1[6].value, ld1[7].value, ld1[8].value, ld1[9].value,
           ld1[10].value, ld1[11].value])
    sheet.append([ld1[0].value, ld1[1].value, ld1[2].value, ld1[3].value, ld1[4].value,
                  ld1[5].value, ld1[6].value, ld1[7].value, ld1[8].value, ld1[9].value,
                  ld1[10].value, ld1[11].value])

res_wb.save("附件五正异常坐标合并后.xlsx")