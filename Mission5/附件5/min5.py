import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from mp5 import optimize

res_wb = Workbook()
res_ws = res_wb.create_sheet()
sheet = res_wb.active
title = ["Tag编号", "A0训练值", "A1训练值", "A2训练值", "A3训练值", "A0A1A2x", "A0A1A2y", "A0A1A2z", "预测精准度",
         "最优解x", "最优解y", "最优解z"]
sheet.append(title)

wb = openpyxl.load_workbook("根据附件5无干扰训练数据得到坐标.xlsx", data_only=True)
# wb = openpyxl.load_workbook("根据附件5有干扰训练数据得到坐标.xlsx", data_only=True)
ws = wb.active
rows = ws.rows
# 全部数据
li = list(rows)
i=1
error=0
for i in range(1, len(li)):
    train_row = li[i]
    S = [train_row[1].value, train_row[2].value, train_row[3].value, train_row[4].value]
    # print(S)
    G = [train_row[5].value, train_row[6].value, train_row[7].value]
    anchor = [[0, 0, 1300],
              [5000, 0, 1700],
              [0, 5000, 1700],
              [5000, 5000, 1300]]
    # print(S, G)
    ans = optimize(anchor, S, G)
    if ans[0]:
        sheet.append(
            [int(train_row[0].value), S[0], S[1], S[2], S[3], train_row[5].value, train_row[6].value, train_row[7].value,
             ans[0], ans[1][0], ans[1][1], ans[1][2]])
    else:
        error+=1
        red_fill = PatternFill(fill_type='solid', fgColor="FF0000", bgColor="AACF91")
        sheet.row_dimensions[i+1].fill = red_fill
        sheet.append(
            [i, S[0], S[1], S[2], S[3], train_row[5].value, train_row[6].value, train_row[7].value,
             ans[0], ans[1][0], ans[1][1], ans[1][2]])

i += 1
print(error)
# res_wb.save("利用COBYLA算法求得最优解.xlsx")
# 0/5
res_wb.save("利用SLSQP算法求得附件5无干扰数据最优解.xlsx")
# 0/5
# res_wb.save("利用SLSQP算法求得附件5有干扰数据最优解.xlsx")