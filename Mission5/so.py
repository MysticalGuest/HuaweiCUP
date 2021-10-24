import os
import re
from openpyxl import Workbook, load_workbook


def read_every_txt(s, datas):
    # 验证数据是否重复的数组
    diff = []
    # 正常数据数
    num_normal = 0
    # 重复数据
    num_repeat = 0
    # 异常数据
    num_error = 0
    # 缺失数据
    num_defect = 0
    # for i in range(len(datas)):
    i = 0
    while i < len(datas):
        # print(len(datas[i]))
        arr = re.findall(r"\d+", datas[i])
        t = arr[0]
        # print(t+"      ", end='')
        a_list = [arr[3]]
        # a_list.append()
        check_list = [arr[4]]
        check_no = arr[5]
        serial = arr[6]
        # 4个一组
        for j in range(1, 4):
            if i+j > len(datas)-1:
                break
            arr_temp = re.findall(r"\d+", datas[i+j])
            # print(arr_temp)
            if t != arr_temp[0]:
                # i += j
                # num_error += j
                break
            a_list.append(arr_temp[3])
            check_list.append(arr_temp[4])
        if len(a_list) == 4:
            if a_list != check_list:
                i += 4
                print("Tag标识: " + t + "的数据与校验值不同！")
                break
            if a_list not in diff:
                li = [t, int(a_list[0]), int(a_list[1]), int(a_list[2]), int(a_list[3]), check_no, serial]
                if float(min(li[1:5]))/float(max(li[1:5])) > 10:
                    num_error += 1
                    i += 4
                    continue
                num_normal += 1
                diff.append([a_list[0], a_list[1], a_list[2], a_list[3]])
                s.append(li)
            else:
                num_repeat += 4
            i += 4
        else:
            num_defect += len(a_list)
            # print(str(i)+ "         ", end='')
            i += len(a_list)
    print("文件中共有"+str(len(datas))+"条数据，去除"+str(num_defect)+"缺失数据，"
          +str(num_repeat)+"条重复数据，"+str(num_error)+"条异常数据，经处理后，保留了"+str(num_normal)+"个样本")
    return [num_repeat, num_error, num_normal, num_defect]


if __name__ == '__main__':
    all = 0
    all_normal = 0
    all_error = 0
    # 缺失数据
    all_defect = 0
    all_repeat = 0

    with open("附件5：动态轨迹数据.txt", 'r') as txtData:  # 打开文件夹中的文件内容
        lines = txtData.readlines()
        # print(lines[1:])
        wb = Workbook()
        ws = wb.create_sheet()
        sheet = wb.active
        title = ["Tag标识", "A0", "A1", "A2", "A3", "数据序列号", "数据编号"]
        sheet.append(title)
        all += len(lines[1:])
        temp = read_every_txt(sheet, lines[1:])
        all_repeat += temp[0]
        all_error += temp[1]
        all_normal += temp[2]
        all_defect += temp[3]
        # 处理文件名
        wb.save("附件5数据.xlsx")

