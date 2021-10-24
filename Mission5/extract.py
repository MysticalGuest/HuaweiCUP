from openpyxl import load_workbook


def extract_locationX():
    res = []
    wb = load_workbook("利用SLSQP算法求得最优解.xlsx", data_only=True)
    ws = wb.active
    rows = ws.rows
    # 全部数据
    li = list(rows)
    # print(li)
    for ld in li[1:]:
        res.append(ld[5].value)
    print("x数据提取完成！")
    return res


def extract_locationY():
    res = []
    wb = load_workbook("利用SLSQP算法求得最优解.xlsx", data_only=True)
    ws = wb.active
    rows = ws.rows
    # 全部数据
    li = list(rows)
    for ld in li[1:]:
        res.append(ld[6].value)
    print("y数据提取完成！")
    return res


def extract_locationZ():
    res = []
    wb = load_workbook("利用SLSQP算法求得最优解.xlsx", data_only=True)
    ws = wb.active
    rows = ws.rows
    # 全部数据
    li = list(rows)
    for ld in li[1:]:
        res.append(ld[7].value)
    print("z数据提取完成！")
    return res

