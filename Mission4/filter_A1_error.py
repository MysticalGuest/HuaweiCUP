import os
import re
import openpyxl
from openpyxl import Workbook


def filter_data(s, num, lis, av):
    for da in lis:
        # if abs(int(da[1].value)-int(av[1].value)) > 400:
        #     s.append([int(num), int(da[1].value), int(da[2].value), int(da[3].value), int(da[4].value)])
        # if abs(int(da[2].value)-int(av[2].value)) > 400:
        #     s.append([int(num), int(da[1].value), int(da[2].value), int(da[3].value), int(da[4].value)])
        # if abs(int(da[3].value)-int(av[3].value)) > 400:
        #     s.append([int(num), int(da[1].value), int(da[2].value), int(da[3].value), int(da[4].value)])
        if abs(int(da[4].value)-int(av[4].value)) > 400:
            # print(int(da[4].value), int(av[4].value), abs(int(da[4].value)-int(av[4].value)) )
            s.append([int(num), int(da[1].value), int(da[2].value), int(da[3].value), int(da[4].value)])


if __name__ == '__main__':
    # root_dir = "..\\Mission1\\正常数据\\"
    root_dir = "..\\Mission1\\异常数据\\"
    r_dir = "..\\Mission2\\"

    wb_a = openpyxl.load_workbook(r_dir+"全部324条正常数据平均值.xlsx", data_only=True)
    ws_a = wb_a.active
    rows_a = ws_a.rows
    # 全部数据
    li_a = list(rows_a)

    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    # title = ["Tag编号", "A0异常", "A1", "A2", "A3"]
    # title = ["Tag编号", "A0", "A1异常", "A2", "A3"]
    # title = ["Tag编号", "A0", "A1", "A2异常", "A3"]
    title = ["Tag编号", "A0", "A1", "A2", "A3异常"]
    sheet.append(title)

    # i=1
    # for (dirpath, dirnames, filenames) in os.walk(root_dir):
        # print(filenames)
    for i in range(1, 325):
        # print(str(i) + ".异常.xlsx:        ", end='')
        wb = openpyxl.load_workbook(root_dir + str(i) + ".异常.xlsx", data_only=True)
    # for filename in filenames:
    #     # print(filename)
    #     wb = openpyxl.load_workbook(root_dir+filename, data_only=True)
        ws = wb.active
        rows = ws.rows
        # 全部数据
        li = list(rows)
        # no = re.search(r"\d+", filename).group()
        filter_data(sheet, int(i), li[1:], li_a[i])
    i+=1
    # res_wb.save("所有A0异常的数据.xlsx")
    # res_wb.save("所有A1异常的数据.xlsx")
    # res_wb.save("所有A2异常的数据.xlsx")
    res_wb.save("所有A3异常的数据.xlsx")