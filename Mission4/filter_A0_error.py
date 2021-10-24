import os
import re
from math import sqrt

import openpyxl
from openpyxl import Workbook, load_workbook

def filter_data(s, num, lis, av):
    # a = [av[1], av[2], av[3], av[4]]
    for da in lis:
        # print(da[0].value, da[1].value,av[1].value)
        # print(num, da[1].value, da[2].value, da[3].value)
        if int(da[1].value)-int(av[1].value)>400:
            s.append([int(num), int(da[1].value), int(da[2].value), int(da[3].value), int(da[4].value)])





if __name__ == '__main__':
    # root_dir = "..\\Mission1\\正常数据\\"
    root_dir = "..\\Mission1\\异常数据\\"
    r_dir = "..\\Mission2\\"

    wb_a = openpyxl.load_workbook(r_dir+"全部324条正常数据平均值.xlsx", data_only=True)
    ws_a = wb_a.active
    rows_a = ws_a.rows
    # 全部数据
    li_a = list(rows_a)

    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    title = ["Tag编号", "A0异常", "A1", "A2", "A3"]
    sheet.append(title)

    i=1
    for (dirpath, dirnames, filenames) in os.walk(root_dir):
        # print(filenames)
        for filename in filenames:
            wb = openpyxl.load_workbook(root_dir+filename, data_only=True)
            ws = wb.active
            rows = ws.rows
            # 全部数据
            li = list(rows)
            no = re.search("\d+", filename).group()
            filter_data(sheet, no, li[1:], li_a[i])
        i+=1
    res_wb.save("所有A0异常的数据.xlsx")
    # # print(li_a[1][2].value)
    # i=1
    # with open("Tag坐标信息.txt", 'r', encoding='utf-8') as txtData:  # 打开文件夹中的文件内容
    #     lines = txtData.readlines()
    #     # print(lines[2:])
    #     for l in lines[2:]:
    #         print(l, end='')
    #         if l != '\n':
    #
    #
    #             loc = re.findall(r"\d+", l)
    #             # wb = openpyxl.load_workbook(root_dir + str(loc[0]) + ".正常.xlsx", data_only=True)
    #             wb = openpyxl.load_workbook(root_dir + str(loc[0]) + ".异常.xlsx", data_only=True)
    #             ws = wb.active
    #             rows = ws.rows
    #             # 全部数据
    #             li = list(rows)
    #
    #             caculate(sheet, li, loc, li_a[i])
    #             # 处理文件名
    #             # name = re.search(r"\d+", filename).group()
    #             res_wb.save("异常数据减500\\"+str(loc[0])+".异常.xlsx")
    #             i+=1
    #         # break
    # print("数据计算完成！")