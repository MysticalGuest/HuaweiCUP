import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from minimize_point import optimize


root_dir = "..\\"

res_wb = Workbook()
res_ws = res_wb.create_sheet()
sheet = res_wb.active
title = ["Tag编号", "精确x", "精确y", "精确z", "预测精准度", "最优解x", "最优解y", "最优解z", "x差值", "y差值", "z差值"]
sheet.append(title)

wb = openpyxl.load_workbook(root_dir+"根据训练数据得到坐标.xlsx", data_only=True)
ws = wb.active
rows = ws.rows
# 全部数据
li = list(rows)
i=1
error=0
with open(root_dir+"Tag坐标信息.txt", 'r', encoding='utf-8') as txtData:
    lines = txtData.readlines()
    for data in lines[2:]:
        if len(data) != 1:
            loc = re.findall(r"\d+", data)
            train_row = li[i]
            S = [train_row[1].value, train_row[2].value, train_row[3].value, train_row[4].value]
            G = [train_row[5].value, train_row[6].value, train_row[7].value]
            anchor = [[0, 0, 1300],
                      [5000, 0, 1700],
                      [0, 5000, 1700],
                      [5000, 5000, 1300]]
            # print(S, G)
            ans = optimize(anchor, S, G)
            # 1 ['1', '50', '50', '88'] True [543.82663116 530.89997737 931.82941295]
            # 324 ['324', '450', '450', '200'] False [718.35134475 676.19105009  14.30062475]
            # print(i, int(loc[1])*10, int(loc[2])*10, int(loc[3])*10, ans[0], ans[1][0], ans[1][1], ans[1][2],
            #       abs(ans[1][0]-int(loc[1])*10),
            #       abs(ans[1][1] - int(loc[2]) * 10),
            #       abs(ans[1][2] - int(loc[3]) * 10)
            # )
            if ans[0]:
                sheet.append(
                    [i, int(loc[1]) * 10, int(loc[2]) * 10, int(loc[3]) * 10, ans[0], ans[1][0], ans[1][1], ans[1][2],
                     abs(ans[1][0] - int(loc[1]) * 10),
                     abs(ans[1][1] - int(loc[2]) * 10),
                     abs(ans[1][2] - int(loc[3]) * 10)])
            else:
                error+=1
                red_fill = PatternFill(fill_type='solid', fgColor="FF0000", bgColor="AACF91")
                sheet.row_dimensions[i+1].fill = red_fill
                sheet.append([i, int(loc[1])*10, int(loc[2])*10, int(loc[3])*10, ans[0], ans[1][0], ans[1][1], ans[1][2],
                      abs(ans[1][0]-int(loc[1])*10),
                      abs(ans[1][1] - int(loc[2]) * 10),
                      abs(ans[1][2] - int(loc[3]) * 10)])

        i += 1
print(error)
# res_wb.save("利用COBYLA算法求得最优解.xlsx")
res_wb.save("利用SLSQP算法求得最优解.xlsx")
# res_wb.save("利用TNC算法求得最优解.xlsx")
# res_wb.save("利用Nelder-Mead算法求得最优解.xlsx")

# res_wb.save("利用Powell算法求得最优解.xlsx")
# res_wb.save("利用trust-krylov算法求得最优解.xlsx")