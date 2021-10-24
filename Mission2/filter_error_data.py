import re
import openpyxl
# from data_process import extract_num, handle_tag
from openpyxl.styles import colors, PatternFill
import numpy as np
from openpyxl import Workbook, load_workbook


def caculate(s, list_rows, i):
    # 除标题之外的所有行
    next_rows = list_rows[1:]
    sum_A0 = 0
    sum_A1 = 0
    sum_A2 = 0
    sum_A3 = 0
    for row in next_rows:
        sum_A0 += int(row[1].value)
        sum_A1 += int(row[2].value)
        sum_A2 += int(row[3].value)
        sum_A3 += int(row[4].value)
        # print(row[1].value, row[2].value, row[3].value, row[4].value)
    all = len(next_rows)
    print(sum_A0 / all, sum_A1 / all, sum_A2 / all, sum_A3 / all)
    s.append([str(i), sum_A0 / all, sum_A1 / all, sum_A2 / all, sum_A3 / all])


if __name__ == '__main__':
    root_dir = "..\\Mission1\\异常数据\\"
    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    title = ["Tag编号", "A0平均值", "A1平均值", "A2平均值", "A3平均值"]
    sheet.append(title)
    for i in range(324):
        print(str(i+1) + ".异常.xlsx:        ", end='')
        wb = openpyxl.load_workbook(root_dir + str(i+1) + ".异常.xlsx", data_only=True)
        ws = wb.active
        rows = ws.rows
        # 全部数据
        li = list(rows)
        # 获得第一行
        # first_row = list_rows[0]
        # # print(first_row)
        # for elem in first_row:
        #     print(elem.value)
        caculate(sheet, li, i+1)
    res_wb.save("全部324条异常数据平均值.xlsx")
    print("数据筛选完成！")




