from math import sqrt

from openpyxl import Workbook, load_workbook
from scipy.optimize import fsolve
import numpy as np

s=[]


def solve_function123(unsolved_value):
    x, y, z = unsolved_value[0], unsolved_value[1], unsolved_value[2]
    return [
        x ** 2 + y ** 2 + (z - 1300) ** 2 - int(s[0]) ** 2,
        (x - 5000) ** 2 + y ** 2 + (z - 1700) ** 2 - int(s[1]) ** 2,
        x ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - int(s[2]) ** 2,
        # (x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - s3 ** 2,
    ]


def solve_function124(unsolved_value):
    x, y, z = unsolved_value[0], unsolved_value[1], unsolved_value[2]
    return [
        x ** 2 + y ** 2 + (z - 1300) ** 2 - int(s[0]) ** 2,
        (x - 5000) ** 2 + y ** 2 + (z - 1700) ** 2 - int(s[1]) ** 2,
        (x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - int(s[3]) ** 2,
    ]


def solve_function134(unsolved_value):
    x, y, z = unsolved_value[0], unsolved_value[1], unsolved_value[2]
    return [
        x ** 2 + y ** 2 + (z - 1300) ** 2 - int(s[0]) ** 2,
        x ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - int(s[2]) ** 2,
        (x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - int(s[3]) ** 2,
    ]


def solve_function234(unsolved_value):
    x, y, z = unsolved_value[0], unsolved_value[1], unsolved_value[2]
    return [
        (x - 5000) ** 2 + y ** 2 + (z - 1700) ** 2 - int(s[1]) ** 2,
        x ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - int(s[2]) ** 2,
        (x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2 - int(s[3]) ** 2,
    ]


def judge(data, stand, sheet):
    # print(lis)
    lis = [data[1].value, data[2].value, data[3].value, data[4].value]
    global s
    s = lis
    print(s)
    error = s[0]
    solved = fsolve(solve_function234, [0, 0, 0])

    v = list(map(lambda x: x[0] - x[1], zip(solved, stand)))
    r0 = sqrt(solved[0] ** 2 + solved[1] ** 2 + (solved[2]-1300) ** 2)
    sub0 = r0+500-float(error)

    error = s[1]
    solved = fsolve(solve_function134, [0, 0, 0])
    v = list(map(lambda x: x[0] - x[1], zip(solved, stand)))
    r1 = sqrt((solved[0]-5000) ** 2 + solved[1] ** 2 + (solved[2] - 1700) ** 2)
    sub1 = r1 + 500 - float(error)

    error = s[2]
    solved = fsolve(solve_function124, [0, 0, 0])
    v = list(map(lambda x: x[0] - x[1], zip(solved, stand)))
    r2 = sqrt(solved[0] ** 2 + (solved[1]-5000) ** 2 + (solved[2] - 1700) ** 2)
    sub2 = r2 + 500 - float(error)

    error = s[3]
    solved = fsolve(solve_function123, [0, 0, 0])
    v = list(map(lambda x: x[0] - x[1], zip(solved, stand)))
    r3 = sqrt((solved[0]-5000) ** 2 + (solved[1]-5000) ** 2 + (solved[2] - 1300) ** 2)
    sub3 = r3 + 500 - float(error)

    c = [sub0, sub1, sub2, sub3]
    print(c)
    index = c.index(min(c))
    # print(data[0].value, lis[0], lis[1], lis[2], lis[3], "A"+str(index), data[index+1].value)
    sheet.append([data[0].value, lis[0], lis[1], lis[2], lis[3], "A"+str(index), data[index+1].value])
    # print(index)




# wb = load_workbook("所有异常数据.xlsx", data_only=True)
wb = load_workbook("2.异常.xlsx", data_only=True)
ws = wb.active
# title = ["Tag标识", "A0", "A1", "A2", "A3", "出错数据标签", "出错数据数值"]
# ws.append(title)
rows = ws.rows
# 全部数据
li = list(rows)

res_wb = Workbook()
res_ws = res_wb.create_sheet()
sheet = res_wb.active
title = ["Tag标识", "A0", "A1", "A2", "A3", "出错数据标签", "出错数据数值"]
sheet.append(title)
for ld in li[1:]:
    # for r in ld:
        # print(r.value, end=' ')
    # judge([ld[1].value, ld[2].value, ld[3].value, ld[4].value], [500, 500, 880], ws)
    judge(ld, [500, 500, 880], sheet)
    # print()
# res_wb.save("2.异常筛选完毕.xlsx")