import os
import re

import openpyxl
from openpyxl import Workbook, load_workbook


def caculate(s, list_rows, location):
    # 除标题之外的所有行
    next_rows = list_rows[1:]
    sum_A0 = 0
    sum_A1 = 0
    sum_A2 = 0
    sum_A3 = 0
    for row in next_rows:
        sum_A0 += int(row[1].value)
        sum_A1 += int(row[2].value)
        sum_A2 += int(row[3].value)
        sum_A3 += int(row[4].value)
    all = len(next_rows)
    print(sum_A0 / all, sum_A1 / all, sum_A2 / all, sum_A3 / all,
          int(location[1])*10, int(location[2])*10, int(location[3])*10)
    s.append([location[0], sum_A0 / all, sum_A1 / all, sum_A2 / all, sum_A3 / all,
              int(location[1])*10, int(location[2])*10, int(location[3])*10])


if __name__ == '__main__':
    root_dir = "..\\Mission1\\正常数据\\"
    # root_dir = "..\\Mission1\\异常数据\\"
    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    title = ["Tag编号", "A0平均值", "A1平均值", "A2平均值", "A3平均值", "x(mm)", "y(mm)", "z(mm)"]
    sheet.append(title)
    with open("Tag坐标信息.txt", 'r', encoding='utf-8') as txtData:  # 打开文件夹中的文件内容
        lines = txtData.readlines()
        # print(lines[2:])
        for l in lines[2:]:
            if l != '\n':
                loc = re.findall(r"\d+", l)
                # print(l, end='')
                # print(str(loc[0]) + ".正常.xlsx:        ", end='')
                print(str(loc[0]) + ".异常.xlsx:        ", end='')
                wb = openpyxl.load_workbook(root_dir + str(loc[0]) + ".正常.xlsx", data_only=True)
                # wb = openpyxl.load_workbook(root_dir + str(loc[0]) + ".异常.xlsx", data_only=True)
                ws = wb.active
                rows = ws.rows
                # 全部数据
                li = list(rows)
                caculate(sheet, li, loc)
    res_wb.save("全部324条正常数据平均值(带坐标).xlsx")
    # res_wb.save("全部324条异常数据平均值(带坐标).xlsx")
    print("数据筛选完成！")