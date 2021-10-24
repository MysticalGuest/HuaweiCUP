
import re

import openpyxl
from openpyxl import Workbook


if __name__ == '__main__':
    # root_dir = "..\\Mission1\\正常数据\\"
    root_dir = "..\\..\\Mission1\\异常数据\\"

    wb_a = openpyxl.load_workbook("..\\全部324条正常数据平均值.xlsx", data_only=True)
    ws_a = wb_a.active
    rows_a = ws_a.rows
    # 全部数据
    li_a = list(rows_a)
    # print(li_a[1][2].value)
    i=1
    with open("..\\Tag坐标信息.txt", 'r', encoding='utf-8') as txtData:  # 打开文件夹中的文件内容
        lines = txtData.readlines()
        for l in lines[2:]:
            print(l, end='')
            if l != '\n':
                res_wb = Workbook()
                res_ws = res_wb.create_sheet()
                sheet = res_wb.active
                title = ["Tag编号", "A0", "A1", "A2", "A3"]
                sheet.append(title)

                loc = re.findall(r"\d+", l)
                # wb = openpyxl.load_workbook(root_dir + str(loc[0]) + ".正常.xlsx", data_only=True)
                wb = openpyxl.load_workbook(root_dir + str(loc[0]) + ".异常.xlsx", data_only=True)
                ws = wb.active
                rows = ws.rows
                # 全部数据
                li = list(rows)
                for ld in li[1:]:
                    # if abs(int(ld[1].value)-li_a[i][1].value) > 400:
                    #     sheet.append([ld[0].value, ld[1].value, ld[2].value, ld[3].value, ld[4].value])
                    # if abs(int(ld[2].value)-li_a[i][2].value) > 400:
                    #     sheet.append([ld[0].value, ld[1].value, ld[2].value, ld[3].value, ld[4].value])
                    # if abs(int(ld[3].value)-li_a[i][3].value) > 400:
                    #     sheet.append([ld[0].value, ld[1].value, ld[2].value, ld[3].value, ld[4].value])
                    if abs(int(ld[4].value)-li_a[i][4].value) > 400:
                        sheet.append([ld[0].value, ld[1].value, ld[2].value, ld[3].value, ld[4].value])

                # res_wb.save("异常数据A1\\" + str(loc[0]) + ".异常.xlsx")
                # res_wb.save("异常数据A2\\" + str(loc[0]) + ".异常.xlsx")
                # res_wb.save("异常数据A0\\" + str(loc[0]) + ".异常.xlsx")
                res_wb.save("异常数据A3\\" + str(loc[0]) + ".异常.xlsx")
                i+=1
                # break
            # break
    print("数据筛选完成！")