
import re

import openpyxl
from openpyxl import Workbook


if __name__ == '__main__':
    # root_dir = "..\\Mission1\\正常数据\\"
    # root_dir = "异常数据A0\\"
    # root_dir = "异常数据A1\\"
    # root_dir = "异常数据A2\\"
    root_dir = "异常数据A3\\"
    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    # title = ["Tag编号", "A1", "A2", "A3", "x", "y", "z"]
    # title = ["Tag编号", "A0", "A2", "A3", "x", "y", "z"]
    # title = ["Tag编号", "A0", "A1", "A3", "x", "y", "z"]
    title = ["Tag编号", "A0", "A1", "A2", "x", "y", "z"]
    sheet.append(title)

    i=1
    with open("..\\Tag坐标信息.txt", 'r', encoding='utf-8') as txtData:  # 打开文件夹中的文件内容

        lines = txtData.readlines()
        for l in lines[2:]:
            if i > 324:
                break
            loc = re.findall(r"\d+", l)
            wb = openpyxl.load_workbook(root_dir + str(i) + ".异常.xlsx", data_only=True)
            ws = wb.active
            rows = ws.rows
            # 全部数据
            li = list(rows)
            if len(li) > 1:
                for ld in li[1:]:
                    # sheet.append([i, ld[2].value, ld[3].value, ld[4].value, loc[1], loc[2], loc[3]])
                    # sheet.append([i, ld[1].value, ld[3].value, ld[4].value, loc[1], loc[2], loc[3]])
                    # sheet.append([i, ld[1].value, ld[2].value, ld[4].value, loc[1], loc[2], loc[3]])
                    sheet.append([i, ld[1].value, ld[2].value, ld[3].value, loc[1], loc[2], loc[3]])
            i+=1
    # res_wb.save("去掉异常A0数据后得到正常样本.xlsx")
    # res_wb.save("去掉异常A1数据后得到正常样本.xlsx")
    # res_wb.save("去掉异常A2数据后得到正常样本.xlsx")
    res_wb.save("去掉异常A3数据后得到正常样本.xlsx")
    print("数据筛选完成！")