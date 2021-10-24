import re
import openpyxl
# from data_process import extract_num, handle_tag
from openpyxl.styles import colors, PatternFill
import numpy as np
from openpyxl import Workbook, load_workbook
import random


if __name__ == '__main__':
    root_dir = "..\\..\\Mission1\\异常数据\\"
    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    title = ["Tag编号", "A0", "A1", "A2", "A3"]
    sheet.append(title)
    for i in range(324):
        # print(str(i+1) + ".正常.xlsx:        ", end='')
        wb = openpyxl.load_workbook(root_dir + str(i+1) + ".异常.xlsx", data_only=True)
        ws = wb.active
        rows = ws.rows
        # 全部数据
        li = list(rows)
        select_data = li[random.randint(1, len(li) - 1)]
        # print(select_data)
        sheet.append([int(i+1), int(select_data[1].value), int(select_data[2].value),
                      int(select_data[3].value), int(select_data[4].value)])
    res_wb.save("全部324条异常数据中选一条.xlsx")
    print("完成！")