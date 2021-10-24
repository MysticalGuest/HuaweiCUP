"""
@version: v1.0
@author: 你说得很队 21102860168
@license: Apache Licence
@contact: mysticalguest@163.com
@site: 
@software: PyCharm
@file: average_after_subtract_500.py
@time: 2021/10/21 19:31
"""
import openpyxl
from openpyxl import Workbook


def caculate(s, list_rows, i):
    # 除标题之外的所有行
    next_rows = list_rows[1:]
    sum_A0 = 0
    sum_A1 = 0
    sum_A2 = 0
    sum_A3 = 0
    for row in next_rows:
        sum_A0 += int(row[5].value)
        sum_A1 += int(row[6].value)
        sum_A2 += int(row[7].value)
        sum_A3 += int(row[8].value)
    all = len(next_rows)
    print(sum_A0 / all, sum_A1 / all, sum_A2 / all, sum_A3 / all)
    s.append([str(i), sum_A0 / all, sum_A1 / all, sum_A2 / all, sum_A3 / all])


def average_after_subtract(root_dir, save_name):
    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    title = ["Tag编号", "A0平均值", "A1平均值", "A2平均值", "A3平均值"]
    sheet.append(title)
    for i in range(324):
        wb = openpyxl.load_workbook(root_dir + str(i + 1) + ".异常.xlsx", data_only=True)
        ws = wb.active
        rows = ws.rows
        # 全部数据
        li = list(rows)

        caculate(sheet, li, i + 1)
    res_wb.save(save_name)
    print("数据计算完成！")

