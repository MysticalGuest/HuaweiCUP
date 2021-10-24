import random
import re
from math import sqrt

import openpyxl
import os
from openpyxl import Workbook, load_workbook


root_dir = "..\\Mission1\\正常数据\\"

res_wb = Workbook()
res_ws = res_wb.create_sheet()
sheet = res_wb.active
title = ["Tag编号", "A0", "A1", "A2", "A3", "x", "y", "z"]
sheet.append(title)

loc = []
with open("Tag坐标信息.txt", 'r', encoding='utf-8') as txtData:  # 打开文件夹中的文件内容
    lines = txtData.readlines()
    for lo in lines[2:]:
        no = re.findall(r"\d+", lo)
        if len(no)!=0:
            loc.append(no)
            # print(no)
    # loc = []
# print(loc)
i = 0
for (dirpath, dirnames, filenames) in os.walk(root_dir):
    for filename in filenames:
        no = re.search(r"\d+", filename).group()
        # print(no)
        # 计算精确距离
        # x = float(loc[i][1]) * 10
        # y = float(loc[i][2]) * 10
        # z = float(loc[i][3]) * 10
        # # print(x,y,z)
        # A0 = sqrt(x ** 2 + y ** 2 + (z - 1300) ** 2)
        # A1 = sqrt((x - 5000) ** 2 + y ** 2 + (z - 1700) ** 2)
        # A2 = sqrt(x ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2)
        # A3 = sqrt((x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1300) ** 2)
        # print(filename)
        no = re.search(r"\d+", filename).group()
        wb = openpyxl.load_workbook(root_dir + filename, data_only=True)
        ws = wb.active
        rows = ws.rows
        # 全部数据
        li = list(rows)
        # print(li[1:])
        for row in li[1:]:
            # print(loc[i][0], row[1].value, row[2].value, row[3].value, row[4].value, A0, A1, A2, A3)
            sheet.append([int(no), int(row[1].value), int(row[2].value), int(row[3].value), int(row[4].value),
                          int(loc[i][1])*10, int(loc[i][2])*10, int(loc[i][3])*10])
        i += 1
        if i>323:
            break
            # print(row[1].value)
res_wb.save("所有正常数据.xlsx")
