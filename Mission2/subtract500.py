import re

import openpyxl
from openpyxl import Workbook


def judge(jl, ac):
    for jd, a in zip(jl, ac):
        if jd-a < 100:
            jd -= 500


def caculate(s, list_rows, location, av):
    # 除标题之外的所有行
    next_rows = list_rows[1:]

    for row in next_rows:
        be_A0 = int(row[1].value)
        be_A1 = int(row[2].value)
        be_A2 = int(row[3].value)
        be_A3 = int(row[4].value)
        be = [be_A0, be_A1, be_A2, be_A3]
        a = [av[1], av[2], av[3], av[4]]

        for i in range(4):
            print(be[i], a[i].value)
            if be[i] - a[i].value > 400:
                be[i] -= 500
        s.append([location[0], row[1].value, row[2].value, row[3].value, row[4].value,
                  be[0], be[1], be[2],  be[3]])


if __name__ == '__main__':
    root_dir = "..\\Mission1\\异常数据\\"

    wb_a = openpyxl.load_workbook("全部324条正常数据平均值.xlsx", data_only=True)
    ws_a = wb_a.active
    rows_a = ws_a.rows
    # 全部数据
    li_a = list(rows_a)
    i=1
    with open("Tag坐标信息.txt", 'r', encoding='utf-8') as txtData:  # 打开文件夹中的文件内容
        lines = txtData.readlines()
        # print(lines[2:])
        for l in lines[2:]:
            print(l, end='')
            if l != '\n':
                res_wb = Workbook()
                res_ws = res_wb.create_sheet()
                sheet = res_wb.active
                title = ["Tag编号", "A0", "A1", "A2", "A3", "A0修改后", "A1修改后", "A2修改后", "A3修改后"]
                sheet.append(title)

                loc = re.findall(r"\d+", l)
                # wb = openpyxl.load_workbook(root_dir + str(loc[0]) + ".正常.xlsx", data_only=True)
                wb = openpyxl.load_workbook(root_dir + str(loc[0]) + ".异常.xlsx", data_only=True)
                ws = wb.active
                rows = ws.rows
                # 全部数据
                li = list(rows)

                caculate(sheet, li, loc, li_a[i])
                # 处理文件名
                res_wb.save("异常数据减500\\"+str(loc[0])+".异常.xlsx")
                i+=1
            # break
    print("数据计算完成！")