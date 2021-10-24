"""
@version: v1.0
@author: 你说得很队 21102860168
@license: Apache Licence
@contact: mysticalguest@163.com
@site: 
@software: PyCharm
@file: merge_all_error_data.py
@time: 2021/10/15 08:43
"""

import re
from math import sqrt

import openpyxl
import os
from openpyxl import Workbook


def merge(root_dir, save_name):

    res_wb = Workbook()
    res_ws = res_wb.create_sheet()
    sheet = res_wb.active
    title = ["Tag编号", "A0", "A1", "A2", "A3", "A0精确值", "A1精确值", "A2精确值", "A3精确值"]
    sheet.append(title)

    loc = []
    with open("Tag坐标信息.txt", 'r', encoding='utf-8') as txtData:  # 打开文件夹中的文件内容
        lines = txtData.readlines()
        for lo in lines[2:]:
            no = re.findall(r"\d+", lo)
            if len(no)!=0:
                loc.append(no)
    i = 0
    for (dirpath, dirnames, filenames) in os.walk(root_dir):
        for filename in filenames:
            # 计算精确距离
            x = float(loc[i][1]) * 10
            y = float(loc[i][2]) * 10
            z = float(loc[i][3]) * 10
            A0 = sqrt(x ** 2 + y ** 2 + (z - 1300) ** 2)
            A1 = sqrt((x - 5000) ** 2 + y ** 2 + (z - 1700) ** 2)
            A2 = sqrt(x ** 2 + (y - 5000) ** 2 + (z - 1700) ** 2)
            A3 = sqrt((x - 5000) ** 2 + (y - 5000) ** 2 + (z - 1300) ** 2)

            no = re.search(r"\d+", filename).group()
            wb = openpyxl.load_workbook(root_dir + filename, data_only=True)
            ws = wb.active
            rows = ws.rows
            # 全部数据
            li = list(rows)
            for row in li[1:]:
                sheet.append([int(no), int(row[1].value), int(row[2].value), int(row[3].value), int(row[4].value),
                              A0, A1, A2, A3])
            i += 1
            if i > 323:
                break
    res_wb.save(save_name)
