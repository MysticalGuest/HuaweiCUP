# 每一个文件去一个样本
import random
import re

import openpyxl
# from data_process import extract_num, handle_tag
from openpyxl.styles import colors, PatternFill
import numpy as np
import os
from openpyxl import Workbook, load_workbook


root_dir = "..\\Mission1\\正常数据\\"

res_wb = Workbook()
res_ws = res_wb.create_sheet()
sheet = res_wb.active
title = ["Tag编号", "A0", "A1", "A2", "A3"]
sheet.append(title)
for (dirpath, dirnames, filenames) in os.walk(root_dir):
    for filename in filenames:
        no = re.search(r"\d+", filename).group()
        wb = openpyxl.load_workbook(root_dir+filename, data_only=True)
        ws = wb.active
        rows = ws.rows
        # 全部数据
        li = list(rows)
        select_data = li[random.randint(1, len(li)-1)]
        # print(select_data)
        sheet.append([int(no), int(select_data[1].value), int(select_data[2].value),
                      int(select_data[3].value), int(select_data[4].value)])
    res_wb.save("324条正常数据中取一条.xlsx")
    # print(filenames)
# print(random.randint())